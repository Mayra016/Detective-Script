import random
import openpyxl

# CODEGAME AUTO GUIÓN ES

# Excel with the data base 
wb_obj = openpyxl.load_workbook("/content/drive/MyDrive/Colab Notebooks/Detective.xlsx")
sheet_obj = wb_obj.active


class Subject:
  # Subject properties
  name = ""
  obj = ""
  place = ""
  gender = ""
  isMurder = False
  hasSpecialObjs = False

  # Posible properties
  names = ["Ana", "Jacinta", "Mario", "Juan", "Miguel", "María", "Sandra", "Javier", "Julia", "José", "Pedro", "Angela", "Lucía", "Carolina",
           "Cristina", "Paula", "Carlota"]
  womans = ["Ana", "Jacinta", "María", "Sandra","Julia", "Angela", "Lucía", "Carolina",
           "Cristina", "Paula", "Carlota"]      
  objs = ["anillo", "pendientes", "teléfono", "reloj", "goma de pelo", "goma de mascar", "botella de agua", "pulsera", "gafas", "libro",
          "cuchillo", "barajo", "dados", "gafas de natación", "flauta", "guitarra", "pesas"]   
  femObjs = ["goma de pelo", "goma de mascar", "botella de agua", "pulsera", "gafas", "gafas de natación", "flauta", "guitarra", "pesas"]   
  pluralFemObjs = ["gafas", "gafas de natación", "pesas"]  
  pluralMascObjs = ["pendientes", "dados"]         
  places = ["cocina", "salón", "baño", "biblioteca", "salón de juegos", "sala de música", "piscina", "gimnasio"]
  femPlaces = ["cocina", "biblioteca", "sala de música", "piscina"]
  special_objs = ["cuchillo", "barajo", "dados", "gafas de natación", "flauta", "guitarra", "pesas"]

  def __init__(self):
    self.name = self.getName()
    self.obj = self.generateObj()
    if self.place == "":
      self.place = self.generatePlace()

  def getName(self):
    index = random.randint(0, len(self.names)-1)
    if self.names[index] in self.womans:
      self.gender = "woman"
    else:
      self.gender = "man"  
    return self.names[index]

  def generateObj(self):
    index = random.randint(0, len(self.objs)-1)
    obj = self.objs[index]
    if self.objs[index] in self.special_objs:
      if obj == "cuchillo":
        self.place = "cocina"
      elif obj == "barajo" or obj == "dados":  
        self.place = "salón de juegos"
      elif obj == "gafas de natación":  
        self.place = "piscina"
      elif obj == "flauta" or obj == "guitarra":  
        self.place = "sala de música"  
      elif obj == "pesas":  
        self.place = "gimnasio"  
      #self.place = self.get_specific_place(obj)   
      hasSpecialObjs = True 
    return obj

  def get_specific_place(obj_name):
    if obj_name == "cuchillo":
      return "cocina"
    elif obj_name == "barajo" or obj_name == "dados":  
      return "salón de juegos"
    elif obj_name == "gafas de natación":  
      return "piscina"
    elif obj_name == "flauta" or obj_name == "guitarra":  
      return "sala de música"  
    elif obj_name == "pesas":  
      return "gimnasio"  

  def generatePlace(self):
    index = random.randint(0, len(self.places)-1)  
    return self.places[index] 


def uploadData(subj, currentLevel, clue1, clue2, clue3, clue4, clue5):
    if subj == "subject1":
      cell_name = sheet_obj.cell(row = currentLevel, column = 1);
      cell_name.value = str(subj.name);
      cell_place = sheet_obj.cell(row = currentLevel, column = 4);
      cell_place.value = str(subj.place);
      cell_obj = sheet_obj.cell(row = currentLevel, column = 7);
      cell_obj.value = str(subj.obj);

    if subj == "subject2":
      cell_name = sheet_obj.cell(row = currentLevel, column = 2);
      cell_name.value = str(subj.name);
      cell_place = sheet_obj.cell(row = currentLevel, column = 5);
      cell_place.value = str(subj.place);
      cell_obj = sheet_obj.cell(row = currentLevel, column = 8);
      cell_obj.value = str(subj.obj);  

    if subj == "subject3":
      cell_name = sheet_obj.cell(row = currentLevel, column = 3);
      cell_name.value = str(subj.name);
      cell_place = sheet_obj.cell(row = currentLevel, column = 6);
      cell_place.value = str(subj.place);
      cell_obj = sheet_obj.cell(row = currentLevel, column = 9);
      cell_obj.value = str(subj.obj);
      cell_clue1 = sheet_obj.cell(row = currentLevel, column = 10);
      cell_clue1.value = str(clue1);
      cell_clue2 = sheet_obj.cell(row = currentLevel, column = 11);
      cell_clue2.value = str(clue2)
      cell_clue3 = sheet_obj.cell(row = currentLevel, column = 12);
      cell_clue3.value = str(clue3)
      cell_clue4 = sheet_obj.cell(row = currentLevel, column = 13);
      cell_clue4.value = str(clue4)
      cell_clue5 = sheet_obj.cell(row = currentLevel, column = 14);
      cell_clue5.value = str(clue5)


def tree1(sub1, sub2, sub3):
  subjs = [sub1, sub2, sub3]
  i = 0
  murder = ""
  clue1 = ""
  clue2 = ""
  clue3 = ""
  clue4 = ""
  clue5 = ""

  # Identify the murder
  for sub in subjs:
    if sub.isMurder == True:
      murder = sub

  # Get clue1    
  if murder.obj in murder.femObjs:
    if murder.place in murder.femPlaces:
      clue1 = f'Se encontró una {murder.obj} en la {murder.place} junto al cuerpo'
    else:
      clue1 = f'Se encontró una {murder.obj} en el {murder.place} junto al cuerpo'  
  elif murder.obj in murder.pluralFemObjs:
    if murder.place in murder.femPlaces:
      clue1 = f'Se encontró unas {murder.obj} en la {murder.place} junto al cuerpo'  
    else:  
      clue1 = f'Se encontró unas {murder.obj} en el {murder.place} junto al cuerpo'
  elif murder.obj in murder.pluralMascObjs:
    if murder.place in murder.femPlaces:
      clue1 = f'Se encontró unos {murder.obj} en la {murder.place} junto al cuerpo'  
    else:
      clue1 = f'Se encontró unos {murder.obj} en el {murder.place} junto al cuerpo'    
  else:      
    if murder.place in murder.femPlaces:
      clue1 = f'Se encontró un {murder.obj} en la {murder.place} junto al cuerpo'
    else:
      clue1 = f'Se encontró un {murder.obj} en el {murder.place} junto al cuerpo'  

  # Get clue2
  if murder != sub1:
    person1 = sub1
  else:
    person1 = sub3  
  if person1.obj in person1.femObjs:    
    clue2 = f'{person1.name} tiene una {person1.obj}'
  elif person1.obj in person1.pluralFemObjs:
    clue2 = f'{person1.name} tiene unas {person1.obj}'
  elif person1.obj in person1.pluralMascObjs: 
    clue2 = f'{person1.name} tiene unos {person1.obj}' 
  else:
    clue2 = f'{person1.name} tiene un {person1.obj}' 

  # Get clue3 
  murderOrSub = random.randint(0,1)
  if murder == sub1: 
    if murderOrSub == 0:
      if sub1.place in sub1.femPlaces:    
        clue3 = f'{sub3.name} y {sub2.name} se han visto en la {sub3.place}, pero {sub2.name} salió antes del crimen'
      else:
        clue3 = f'{sub3.name} y {sub2.name} se han visto en el {sub3.place}, pero {sub2.name} salió antes del crimen'
    else:
      if sub1.place in murder.femPlaces:    
        clue3 = f'{sub3.name} y {murder.name} se han visto en la {sub3.place}, pero {murder.name} salió antes del crimen'
      else:
        clue3 = f'{sub3.name} y {murder.name} se han visto en el {sub3.place}, pero {murder.name} salió antes del crimen' 

  elif murder == sub2:
    if murderOrSub == 0:
      if sub1.place in sub1.femPlaces:    
        clue3 = f'{sub1.name} y {sub3.name} se han visto en la {sub1.place}, pero {sub3.name} salió antes del crimen'
      else:
        clue3 = f'{sub1.name} y {sub3.name} se han visto en el {sub1.place}, pero {sub3.name} salió antes del crimen'
    else:
      if sub1.place in murder.femPlaces:    
        clue3 = f'{sub1.name} y {murder.name} se han visto en la {sub1.place}, pero {murder.name} salió antes del crimen'
      else:
        clue3 = f'{sub1.name} y {murder.name} se han visto en el {sub1.place}, pero {murder.name} salió antes del crimen' 

  elif murder == sub3:
    if murderOrSub == 0:
      if sub1.place in sub1.femPlaces:    
        clue3 = f'{sub1.name} y {sub2.name} se han visto en la {sub1.place}, pero {sub2.name} salió antes del crimen'
      else:
        clue3 = f'{sub1.name} y {sub2.name} se han visto en el {sub1.place}, pero {sub2.name} salió antes del crimen'
    else:
      if sub1.place in murder.femPlaces:    
        clue3 = f'{sub1.name} y {murder.name} se han visto en la {sub1.place}, pero {murder.name} salió antes del crimen'
      else:
        clue3 = f'{sub1.name} y {murder.name} se han visto en el {sub1.place}, pero {murder.name} salió antes del crimen'       

  
  # Get clue4
  """
  El accesorio del sujeto2 que indica su localización
  """
  if murder != sub2:
    person = sub2
  else:
    person = sub3  
  if person.obj in person.femObjs and person.place in person.femPlaces:
    clue4 = f'Hay una {person.obj} en la {person.place}'
  elif person.obj in person.femObjs and person.place not in person.femPlaces:  
    clue4 = f'Hay una {person.obj} en el {person.place}'  
  elif person.obj in person.pluralFemObjs and person.place not in person.femPlaces:  
    clue4 = f'Hay unas {person.obj} en el {person.place}'
  elif person.obj in person.pluralFemObjs and person.place in person.femPlaces:
    clue4 = f'Hay unas {person.obj} en la {person.place}'   
  elif person.obj in person.pluralMascObjs and person.place in person.femPlaces:
    clue4 = f'Hay unos {person.obj} en la {person.place}' 
  elif person.obj in person.pluralMascObjs and person.place not in person.femPlaces:
    clue4 = f'Hay unos {person.obj} en el {person.place}' 
  else:
    if person.place in person.femPlaces:
      clue4 = f'Hay un {person.obj} en la {person.place}'
    else:
      clue4 = f'Hay un {person.obj} en el {person.place}'     


  # Get clue5
  """
  El accesorio del Sujeto2 
  """
  if murder != sub2:
    person2 = sub2
  else:
    person2 = sub3  
  if person2.obj in person2.femObjs:
    clue5 = f'{person2.name} tiene una {person2.obj}'
  elif sub2.obj in sub2.pluralFemObjs: 
    clue5 = f'{person2.name} tiene unas {person2.obj}'   
  elif person2.obj in person2.pluralMascObjs:
    clue5 = f'{person2.name} tiene unos {person2.obj}' 
  else:
    clue5 = f'{person2.name} tiene un {person2.obj}'

  return clue1, clue2, clue3, clue4, clue5  
    

def chooseMurder(sub1, sub2, sub3):
  index = random.randint(0, 2)
  murder = ""
  if index == 0:
   sub1.isMurder = True
   murder = sub1
  elif index == 1:
   sub2.isMurder = True 
   murder = sub2
  elif index == 2:
   sub3.isMurder = True 
   murder = sub3
  if murder.hasSpecialObjs == True:
    while murder.obj in murder.special_objs:
      murder.obj = murder.objs[index]
      index = index + 1

def chooseTree(sub1, sub2, sub3):
  index = random.randint(0,4) 
  if index == 1:
    return tree1(sub1, sub2, sub3)      

def removeDuplicated(subject1, subject2, subject3):
  # Duplicated objects
  usedObjs = [subject1.obj, subject2.obj, subject3.obj]
  unchoosedObjs = list(set(subject1.objs).symmetric_difference(usedObjs))
  index = random.randint(0, len(unchoosedObjs))
  if subject1.obj == subject2.obj or subject2.obj == subject3.obj:
    subject2.obj = unchoosedObjs[index]
  elif subject1.obj == subject3.obj:
    if index > 0:
      index = index - 1
    else:
      index = index + 2  
    subject3.obj = unchoosedObjs[index]

  # Duplicated places
  usedPlaces = [subject1.place, subject2.place, subject3.place]
  unchoosedPlaces = list(set(subject1.places).symmetric_difference(usedPlaces))
  indexP = random.randint(0, len(unchoosedPlaces))
  if subject1.place == subject2.place or subject2.place == subject3.place:
    subject2.place = unchoosedPlaces[indexP]
  elif subject1.place == subject3.place:
    if indexP > 0:
      indexP = indexP - 1
    else:
      indexP = indexP + 2  
    subject3.place = unchoosedPlaces[indexP]  

def writeLevels(initial, final):
  currentLevel = initial
  while currentLevel <= final:
    subject1 = Subject()
    subject2 = Subject()
    subject3 = Subject()
    removeDuplicated(subject1, subject2, subject3)
    chooseMurder(subject1, subject2, subject3)
    clue1, clue2, clue3, clue4, clue5 = tree1(subject1, subject2, subject3)
    uploadData(subject1, currentLevel, clue1, clue2, clue3, clue4, clue5)
    uploadData(subject2, currentLevel, clue1, clue2, clue3, clue4, clue5)
    uploadData(subject3, currentLevel, clue1, clue2, clue3, clue4, clue5)
    currentLevel = currentLevel + 1
  wb_obj.save("/content/drive/MyDrive/Colab Notebooks/Detective.xlsx")
